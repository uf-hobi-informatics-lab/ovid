from openpyxl import load_workbook
import sys



pconrnet_cdm_location = "/app/common/cdm/pcornet_7_0/"
pcornet_cdm_file_name = "pcornet_cdm_v70.csv"
pcornet_valueset_file_name = "2025_01_23_PCORnet_Common_Data_Model_v7dot0_parseable.xlsx"


pcornet_cdm_file_path = pconrnet_cdm_location + pcornet_cdm_file_name
# pcornet_cdm_data = pd.read_csv(pcornet_cdm_file_path)


class PcornetCDM:


    

       


###################################################################################################################################
# This function will return the list of fields from the PCORnet CDM based on the passed table name
###################################################################################################################################

   

    @classmethod
    def get_cdm_valueset_df(cls,field_name):
            


            wb = load_workbook(pconrnet_cdm_location+pcornet_valueset_file_name)
            ws = wb["VALUESETS"]
            header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}


            # if "Field name" in header and "valueset items" in header:
            field_col = header["FIELD_NAME"]
            valueset_col = header["VALUESET_ITEM"]
            valueset_list = [
                row[valueset_col - 1]
                for row in ws.iter_rows(min_row=2, values_only=True)
                if row[field_col - 1] == field_name and row[valueset_col - 1] is not None
            ]

            wb.close()

            return  valueset_list




###################################################################################################################################
# This function will 
###################################################################################################################################

    @classmethod
    def get_table_names_contain_a_field_name(cls,field_name):
           

            wb = load_workbook(pconrnet_cdm_location + pcornet_valueset_file_name)
            ws = wb["FIELDS"]

            # Read header to get column indexes
            header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

            # Get the column indices for FIELD_NAME and TABLE_NAME
            field_name_col = header.get("FIELD_NAME")
            table_name_col = header.get("TABLE_NAME")

            # Collect tables where FIELD_NAME == 'PATID'
            tables_with_field_name = set()

            for row in ws.iter_rows(min_row=2, values_only=True):
                if field_name in row[field_name_col - 1] :
                    tables_with_field_name.add(row[table_name_col - 1])

            # Convert to list if needed
            tables_with_field_name = list(tables_with_field_name)

            return tables_with_field_name

            

   
###################################################################################################################################
# This function will 
###################################################################################################################################


            
    @classmethod
    def get_cdm_df(cls, spark):
            


            cdm_table_df = spark.read.load(pcornet_cdm_file_path,format="csv", sep=",", inferSchema="false", header="true",  quote= '"')
       

            return cdm_table_df



# This function will 
###################################################################################################################################


            
    @classmethod
    def get_table_headers(cls, spark, table_name):
            


            cdm_table_df = spark.read.load(pcornet_cdm_file_path,format="csv", sep=",", inferSchema="false", header="true",  quote= '"')

            table_definition = cdm_table_df.filter(cdm_table_df['TABLE_NAME']==table_name)
       
            table_header = table_definition.select("FIELD_NAME").rdd.flatMap(lambda x: x).collect()


            return table_header











