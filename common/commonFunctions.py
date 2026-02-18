
###################################################################################################################################
# This class will host all the common functions used in the mappings by diffrenet mapping scripts
###################################################################################################################################



from itertools import chain
from Crypto.Hash import SHA
from Crypto.Cipher import XOR
from base64 import b64decode, b64encode
from pyspark.sql.functions import udf, col, count, desc, upper, round,  lit, length, max, date_format, Column, to_date
from pyspark.sql import SparkSession 
from pyspark import SparkConf, SparkContext
import pyspark
import time
import subprocess 
from pyspark.sql.types import StructType, StructField, StringType, NullType, DateType, FloatType, TimestampType, DecimalType, DoubleType
import sys
from ovid_secrets import * 
from settings import *
from pcornet_cdm import PcornetCDM
from pyspark.conf import SparkConf
import pandas as pd
from urllib import parse
from datetime import datetime
import pytz
from dateutil import parser as dp
import pyspark.sql.functions as F
import os
import importlib
import openpyxl
from PIL import Image
import re
import snowflake.connector
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import serialization
import glob
import shutil
from openpyxl.utils.exceptions import IllegalCharacterError
from functools import reduce
from openpyxl import load_workbook




class CommonFuncitons:


    def __init__(self, partner):




        self.partner = partner # specifying the partner / used for the hashing
        self.get_time_from_datetime_udf = udf(self.get_time_from_datetime, StringType())
        self.get_date_from_datetime_udf = udf(self.get_date_from_datetime, StringType())
        self.format_date_udf = udf(self.format_date, StringType())
        self.format_time_udf = udf(self.format_time, StringType())
        self.encrypt_id_udf = udf(self.encrypt_id, StringType())
        self.get_current_time_udf = udf(self.get_current_time, StringType())
        self.get_date_from_date_str_with_default_value_udf = udf(self.get_date_from_date_str_with_default_value, StringType())
        self.get_datetime_from_date_str_with_default_value_udf = udf(self.get_datetime_from_date_str_with_default_value, StringType())
        self.copy_if_float_udf= udf(self.copy_if_float, StringType())
        self.return_if_float_udf = udf(self.return_if_float, StringType())
        self.get_valid_zip5_udf = udf(self.get_valid_zip5, StringType())
  






    

    


###################################################################################################################################
# This function will return the xor hashed value based on the partner name and input value/id
###################################################################################################################################

    def xor_encrypt_exact(cls, data, key):
        # --- Match PyCryptodome behavior exactly ---

        # 1. Convert key to bytes
        if not isinstance(key, (bytes, bytearray)):
            key = bytes(key)

        # 2. Truncate key to 32 bytes (PyCryptodome requirement)
        key = key[:32]

        # 3. Convert data to bytes
        if not isinstance(data, (bytes, bytearray)):
            data = data.encode()

        # 4. XOR each byte (identical to underlying C extension)
        key_len = len(key)
        return bytes([data[i] ^ key[i % key_len] for i in range(len(data))])
    

    # @classmethod
    def encrypt_id(cls, id):

            if id is None:

                return None
            else: 


                partner_encryption_value_dict = {

                        'AVH':'AVH',
                        'BND':'BND',
                        'CHP':'CHP',
                        'CHT':'CHT',
                        'EMY':'EMY',
                        'FLM':'FLM',
                        'NCH':'NCH',
                        'ORH':'ORH',
                        'ORL':'ORL',
                        'TMH':'TMH',
                        'TMA':'TMA',
                        'TMC':'TMC',
                        'UAB':'UAB',
                        'UFH':'UFH',
                        'UMI':'UMI',
                        'USF':'USF',
                        'TGH':'TGH',
                        'AMS':'AMS',
                        'UCI':'UCI',
                        'UMN':'UMN',
                        'omop_partner_1'     :'omop_partner_1',
                        'omop_partner_plus'  :'omop_partner_plus',
                        'pcornert_partner_1' :'pcornert_partner_1',
                        'suncoast'           :'suncoast',
                        'NURSING_UFH':'UFH',
                        'NURSING_AVH':'AVH',
                        'TUMOR_UFH':'UFH',


                }


                encryption_seed_file_path = "/ovid_credentials/" + SEED_FILE_NAME

                with open(encryption_seed_file_path, "r") as f:
                    SEED = f.read().strip()


                partner_encryption_value = partner_encryption_value_dict.get(cls.partner.upper(), None)

                # HASHING_KEY = SHA.new(SEED.encode('utf-8')).digest()

                # cipher = XOR.new(HASHING_KEY)

                # return b64encode(cipher.encrypt('{}{}'.format(partner_encryption_value, id))).decode()


                HASHING_KEY = SHA.new(SEED.encode('utf-8')).digest()

                plaintext = f"{partner_encryption_value}{id}".encode()

                encrypted = cls.xor_encrypt_exact(plaintext, HASHING_KEY)

                return b64encode(encrypted).decode()

###################################################################################################################################
# This function will return the current datetime
###################################################################################################################################

   

    @classmethod
    def get_current_time(cls):


        current_utc_datetime = datetime.utcnow()
        new_york_tz = pytz.timezone('America/New_York')

        current_ny_datetime = current_utc_datetime.replace(tzinfo=pytz.utc).astimezone(new_york_tz)

        return current_ny_datetime.strftime("%Y-%m-%d %H:%M:%S")



###################################################################################################################################
# This function will merge pyspark dataframe into one
###################################################################################################################################

   
    @classmethod
    def awk_merge(cls,input_directory, output_directory, prefix, output_name ):

        output_file_path = output_directory +  output_name
        # Update the awk command to use the correct Python variable for the output path
        command = f"awk 'FNR==1 && NR!=1{{next}} 1' {input_directory}/{prefix}*.csv > {output_file_path}"
        subprocess.run(command, shell=True, check=True)


###################################################################################################################################
# This function will remove a directory
###################################################################################################################################

   
    @classmethod
    def delete_directory(cls,directory ):

        command = ["rm","-r", directory]
        # Execute the command
        subprocess.run(" ".join(command), shell=True)

###################################################################################################################################
# This function will only return a valid zip 5 or a null 
###################################################################################################################################

    @classmethod

    def get_valid_zip5(cls, zip5):
        """
        Returns the zip code if it's a valid 5-digit string, otherwise returns None.
        """
        if isinstance(zip5, str) and re.fullmatch(r"\d{5}", zip5):
            return zip5
        return None 
###################################################################################################################################
# This function will output a pyspark dataframe and combine all the sub files into one file and delete the temprary files
###################################################################################################################################

   
    @classmethod
    def write_pyspark_output_file(cls,payspark_df, output_file_name, output_data_folder_path ):

        work_directory = output_data_folder_path+"tmp_"+output_file_name

        payspark_df = payspark_df.select([
            F.lit(None).cast('string').alias(i.name)
            if isinstance(i.dataType, NullType)
            else i.name
            for i in payspark_df.schema
            ])


        payspark_df.write.format("csv").option("header", True).option("nullValue",None).option("delimiter", "\t").option("quote", "").mode("overwrite").save(work_directory)



        output_file_path = os.path.join(output_data_folder_path, output_file_name).replace('tmp_','')
        
        # Update the awk command to use the correct Python variable for the output path
        command = f"awk 'FNR==1 && NR!=1{{next}} 1' {work_directory}/part-*.csv > {output_file_path}"
        subprocess.run(command, shell=True, check=True)

        command = ["rm","-r", work_directory]
        # Execute the command
        subprocess.run(" ".join(command), shell=True)


###################################################################################################################################
# This function will 
###################################################################################################################################

           
    def copy_data(cls,input_file_name,output_file_name,data_folder_path,output_data_folder_path ):

        if not os.path.exists(output_data_folder_path):
            os.makedirs(output_data_folder_path)

        command = ["cp", "-p", os.path.join(data_folder_path, input_file_name), os.path.join(output_data_folder_path,output_file_name)]
                # Execute the command
        subprocess.run(" ".join(command), shell=True)



###################################################################################################################################
# This function will retun the zip to ruca dictionary
###################################################################################################################################

           
    def get_ruca_zip_mapping(cls):

        zip_to_ruca_file_path =  '/app/common/lists/RUCA-codes-2020-zipcode.xlsx'

        zip_to_ruca_file_wb = load_workbook(zip_to_ruca_file_path, data_only=True)

        ruca_2020_zip_code_data_ws = zip_to_ruca_file_wb["RUCA 2020 ZIP Code Data"]
        
        zip_to_ruca = {}



        header = [cell.value for cell in ruca_2020_zip_code_data_ws[2]]

        zip_idx = header.index("ZIPCode") + 1

        ruca_idx = header.index("PrimaryRUCA") + 1


        for row in ruca_2020_zip_code_data_ws.iter_rows(min_row=3, values_only=False):

            zipcode = row[zip_idx - 1].value

            primary_ruca = row[ruca_idx - 1].value

            if zipcode is not None:
                
                zip_to_ruca[str(zipcode).zfill(5)] = primary_ruca

        return zip_to_ruca


###################################################################################################################################
# This function will retun the zip to ruca dictionary
###################################################################################################################################

           
    def get_countyFips_from_county(cls):

        county_fips_path =  '/app/common/lists/RUCA-codes-2020-tract.xlsx'

        zcounty_fips_file_wb = load_workbook(county_fips_path, data_only=True)

        county_fips_ws = zcounty_fips_file_wb["RUCA2020 Tract Data"]
        
        county_to_fips = {}



        header = [cell.value for cell in county_fips_ws[2]]

        county_idx = header.index("CountyName20") + 1

        countyFIPS20_idx = header.index("CountyFIPS20") + 1


        for row in county_fips_ws.iter_rows(min_row=3, values_only=False):

            county = row[county_idx - 1].value

            countyFIPS20 = row[countyFIPS20_idx - 1].value

            if county is not None:
                
                county_to_fips[county.replace(" County","").upper()] = countyFIPS20

        return county_to_fips


###################################################################################################################################
# This function will retun the zip to ruca dictionary
###################################################################################################################################

           
    def get_state_fips_from_zip(cls):



        # getting state  to fips  
        
        state_fips_path =  '/app/common/lists/RUCA-codes-2020-tract.xlsx'

        state_fips_wb = load_workbook(state_fips_path, data_only=True)

        state_fips_ws = state_fips_wb["RUCA2020 Tract Data"]
        
        state_to_fips = {}



        header = [cell.value for cell in state_fips_ws[2]]

        UrbanAreaName20_idx = header.index("UrbanAreaName20") + 1

        state_fips_idx = header.index("StateFIPS20") + 1


        for row in state_fips_ws.iter_rows(min_row=3, values_only=False):

            urban_area_name = row[UrbanAreaName20_idx - 1].value
            
            state_fps = row[state_fips_idx - 1].value

            if urban_area_name and isinstance(urban_area_name, str) and len(urban_area_name) >= 2:
                
                state = urban_area_name[-2:]

                state_to_fips[state] = state_fps





        # getting zip   to state 


        zip_to_ruca_file_path =  '/app/common/lists/RUCA-codes-2020-zipcode.xlsx'

        zip_to_ruca_file_wb = load_workbook(zip_to_ruca_file_path, data_only=True)

        ruca_2020_zip_code_data_ws = zip_to_ruca_file_wb["RUCA 2020 ZIP Code Data"]
        
        zip_to_state = {}



        header = [cell.value for cell in ruca_2020_zip_code_data_ws[2]]

        zip_idx = header.index("ZIPCode") + 1

        state_idx = header.index("State") + 1


        for row in ruca_2020_zip_code_data_ws.iter_rows(min_row=3, values_only=False):

            zipcode = row[zip_idx - 1].value

            state = row[state_idx - 1].value

            if zipcode is not None:
                
                zip_to_state[str(zipcode).zfill(5)] = state



        zip_to_fips = {

            z: state_to_fips.get(s)
            for z, s in zip_to_state.items()
        }



        return zip_to_fips

###################################################################################################################################
# This function will output the mapping gap for the passed formatted file name
###################################################################################################################################

   
    @classmethod
    def deduplicate(cls,partner,file_name,table_name, file_path,folder_name):

        DO_NOT_DEDUPLICATE_LIST = ['TUMOR_DATE']

        spark = cls.get_spark_session('generate mapping report')

        deduplicated_files_path = file_path.replace('formatter', 'deduplicator')
        duplicates_file_path = os.path.join(deduplicated_files_path, 'duplicate_values')
        duplicates_file_name  = file_name.lower()+'_duplicates.csv'
        deduplicated_file_name  = 'deduplicated_'+file_name.lower()+'.csv'

        # Create the deduplicated files directory if it doesn't exist
        if not os.path.exists(deduplicated_files_path):
            os.makedirs(deduplicated_files_path)

        if not os.path.exists(duplicates_file_path):
            os.makedirs(duplicates_file_path)

        input_file = f"{file_path}formatted_{file_name.lower()}.csv"

        df = cls.spark_read(input_file, spark)
        
        row_count = df.count()

        if file_name in DO_NOT_DEDUPLICATE_LIST:

            cls.write_pyspark_output_file(df,deduplicated_file_name, deduplicated_files_path )
            spark.stop()
            return None

        if row_count > AWK_PASSBY_TRESHHOLD :

            spark = cls.get_spark_session('generate mapping report')

            try: 

                if file_name == 'ENROLLMENT':
                                
                    columns_to_consider = [df.columns[0], df.columns[1], df.columns[4]]

                            
                elif file_name == 'DEATH_CAUSE':
                                
                    columns_to_consider = [df.columns[0], df.columns[1],df.columns[2],df.columns[3], df.columns[4]]

           
                elif file_name == 'PAT_RELATIONSHIP':
                                
                    columns_to_consider = [df.columns[0], df.columns[1]]

                elif file_name == 'PCORNET_TRIAL':
                                
                    columns_to_consider = [df.columns[0], df.columns[1], df.columns[2]]


                elif file_name == 'FACT_RELATIONSHIP':
                                
                    columns_to_consider = [df.columns[0], df.columns[1],df.columns[2],df.columns[3], df.columns[4]]

                elif file_name == 'PROBLEM_CONTACT':
                                
                    columns_to_consider = [df.columns[0], df.columns[1],df.columns[2],df.columns[3], df.columns[4], df.columns[5]]
      
                elif file_name == 'INTERVENTION_CONTACT' :
                                
                    columns_to_consider = [df.columns[0], df.columns[1],df.columns[2],df.columns[3], df.columns[4], df.columns[5], df.columns[6]]
   
                      
                    

                else:
                    columns_to_consider = [df.columns[0]]



                deduplicated_df = df.dropDuplicates(columns_to_consider)

                duplicates_df = df.groupBy(*columns_to_consider).agg(count("*").alias("count")).filter("count > 1")

                cls.write_pyspark_output_file(deduplicated_df,deduplicated_file_name, deduplicated_files_path )
                cls.write_pyspark_output_file(duplicates_df,duplicates_file_name, duplicates_file_path )

                spark.stop()

            except Exception as e:

                spark.stop()
                
                cls.print_with_style(str(e), 'danger red')



        else:

            # Set up the deduplicated files path

            spark.stop()

            deduplicated_output_file = f"{deduplicated_files_path}deduplicated_{file_name.lower()}.csv"
            duplicates_output_file = f"{duplicates_file_path}/{file_name.lower()}_duplicates.csv"



            # Construct the gawk script based on the file_name
            if file_name == 'ENROLLMENT':
                gawk_script = f'''
                BEGIN {{ FS = OFS = "\t" }}
                NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
                !seen[$1, $2, $5]++ {{ print > "{deduplicated_output_file}" }}
                seen[$1, $2, $5] > 1 {{ print > "{duplicates_output_file}" }}
                '''
            
            elif file_name == 'DEATH_CAUSE':
                gawk_script = f'''
                BEGIN {{ FS = OFS = "\t" }}
                NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
                !seen[$1, $2, $3, $4, $5]++ {{ print > "{deduplicated_output_file}" }}
                seen[$1, $2, $3, $4, $5] > 1 {{ print > "{duplicates_output_file}" }}
                '''

            elif file_name == 'PAT_RELATIONSHIP':
                gawk_script = f'''
                BEGIN {{ FS = OFS = "\t" }}
                NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
                !seen[$1, $2]++ {{ print > "{deduplicated_output_file}" }}
                seen[$1, $2] > 1 {{ print > "{duplicates_output_file}" }}
                '''

            elif file_name == 'PCORNET_TRIAL':
                gawk_script = f'''
                BEGIN {{ FS = OFS = "\t" }}
                NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
                !seen[$1, $2, $3]++ {{ print > "{deduplicated_output_file}" }}
                seen[$1, $2, $3] > 1 {{ print > "{duplicates_output_file}" }}
                '''


            elif file_name == 'FACT_RELATIONSHIP':
                gawk_script = f'''
                BEGIN {{ FS = OFS = "\t" }}
                NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
                !seen[$1, $2, $3, $4, $5]++ {{ print > "{deduplicated_output_file}" }}
                seen[$1, $2, $3, $4, $5] > 1 {{ print > "{duplicates_output_file}" }}
                '''
            elif file_name == 'PROBLEM_CONTACT':
                gawk_script = f'''
                BEGIN {{ FS = OFS = "\t" }}
                NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
                !seen[$1, $2, $3, $4, $5, $6]++ {{ print > "{deduplicated_output_file}" }}
                seen[$1, $2, $3, $4, $5, $6] > 1 {{ print > "{duplicates_output_file}" }}
                '''
            elif file_name == 'INTERVENTION_CONTACT' :
                gawk_script = f'''
                BEGIN {{ FS = OFS = "\t" }}
                NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
                !seen[$1, $2, $3, $4, $5, $6, $7]++ {{ print > "{deduplicated_output_file}" }}
                seen[$1, $2, $3, $4, $5, $6, $7] > 1 {{ print > "{duplicates_output_file}" }}
                '''


            else:
                gawk_script = f'''
                BEGIN {{ FS = OFS = "\t" }}
                NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
                !seen[$1]++ {{ print > "{deduplicated_output_file}" }}
                seen[$1] > 1 {{ print > "{duplicates_output_file}" }}
                '''

            # Construct the full shell command
            command = f"awk '{gawk_script}' {input_file}"

            subprocess.run(command, shell=True, check=True)




###################################################################################################################################
# This function will 
###################################################################################################################################

    @classmethod
    def dv_match(cls,partner_name,folder,dv_mapping_table_path,fixed_table_path,table_name,output_path,dv_mapper_output_path ):

            spark = cls.get_spark_session('DV MATCH')


            if table_name != 'PROVIDER':

            

                try:

                    fixed_table      = cls.spark_read(fixed_table_path, spark)
                    dv_mapping_table = cls.spark_read(dv_mapping_table_path, spark)


                    joind_df = fixed_table.join(dv_mapping_table, dv_mapping_table["ENCRYPTED_PATID"] == fixed_table['PATID'], how = 'inner')

                    joind_df.show()

                except Exception as e:
                    cls.print_with_style(str(e), 'danger red')

                finally:
                    spark.stop()
            else:

                spark.stop()





###################################################################################################################################
# This function will 
###################################################################################################################################

    @classmethod

    def datavant_transfer_tokens_to_pcornet(cls,input_file_path, file_name, output_file_path,folder,partner):

        datavant_working_directory = f'/app/partners/{partner}/data/datavant_output/{folder}/'

        if os.path.exists(datavant_working_directory):
            shutil.rmtree(datavant_working_directory)

        os.makedirs(datavant_working_directory)



        matching_files = glob.glob(input_file_path+file_name, recursive=True)
        file_names_only = [os.path.basename(f) for f in matching_files]


        partner_settings_path = "partners."+partner+".partner_settings"
        partner_settings = importlib.import_module(partner_settings_path)


        for file_name in file_names_only:


            with open("/ovid_credentials/" +  DV_CREDENTIALS_FILE_NAME, "rb") as cred_file:
                process = subprocess.Popen(
                    [ 
                        "/ovid_credentials/" + DV_TRANSFORMATION_SCRIPT_FILE_NAME,
                        "transform-tokens",
                        "--to", "pcornet",
                        "--from", partner_settings.DATAVANT_KEY,
                        "-s", "oneflorida_crn",
                        "-i", input_file_path + file_name,
                        "-o", datavant_working_directory
                    ],
                    stdin=cred_file,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE
                )
                stdout, stderr = process.communicate()

###################################################################################################################################
# This function will 
###################################################################################################################################

    @classmethod

    def apply_fix(cls,fix_type, fix_name, version, input_path, output_path,fixed_table_name,mapped_table_name,input_data_folder,partner_name, src1, src2):

        if fix_type == 'custom':

            partner_fix_path = f"/app/partners/{partner_name.lower()}/custom_fixes/{fix_name}/{version}.py"


        if fix_type == 'common':



            partner_fix_path = f"/app/common/common_fixes/{fix_name}/{version}.py"

        command = ["python", partner_fix_path, '-f', input_data_folder, '-t', mapped_table_name , '-p', partner_name,'-ftm',fixed_table_name, '-i',input_path,'-o',output_path, '-sr1', src1, '-sr2', src2]

        subprocess.run(" ".join(command), shell=True)

###################################################################################################################################
# This function will output the mapping gap for the passed formatted file name
###################################################################################################################################

   

    @classmethod
    def get_mapping_gap(cls,partner,file_name,table_name, file_path,folder_name):

        
        try: 

            complete_file_path = file_path+'/formatted_'+file_name.lower()+'.csv'
            partner_dictionaries_path = "partners."+partner+".dictionaries"
            partner_dictionaries = importlib.import_module(partner_dictionaries_path)
            
            master_url = os.environ["SPARK_URL"].strip().replace('"', '')
            spark = SparkSession.builder.master(master_url).appName("Ovid").getOrCreate()
            # spark = SparkSession.builder.master("spark://spark-master:7077").appName("MAPPING GAP").getOrCreate()

            df = cls.spark_read(complete_file_path, spark)
            df.cache()
            row_count= df.count()

            mapping_dicionaries_list = []

            for key, value in partner_dictionaries.__dict__.items():
                

                if  ((file_name.lower() in key and 'death_cause' not in key) or  (file_name.lower() in key and file_name.lower()=='death_cause') ) and "_dict" in key:
                    
                    # get the field name


                    prefix = '@' #logic added to only remove the first word from the dictinary name

                    temp_key = prefix+key


                    field_name = temp_key.replace(prefix+file_name.lower()+'_',"").replace('_dict',"").upper()

                    

                    # get the aggreated data from the formatter for the field 

                    # try:

                    aggregated_values = cls.get_aggregate_value_from_table(df,row_count,field_name)


                    # get and output the aggreated data that has mapping 

                
                    dictionary_dataframe = spark.createDataFrame(value.items(), ["dict_key", "dict_value"])
            

                    if 'UNIT' in field_name:
                        joined_df = aggregated_values.join(dictionary_dataframe, dictionary_dataframe['dict_key']==aggregated_values[field_name], how ="left").orderBy(desc("count"))
                    else:
                        joined_df = aggregated_values.join(dictionary_dataframe, upper(dictionary_dataframe['dict_key'])==upper(aggregated_values[field_name]), how ="left").orderBy(desc("count"))
                

                    cls.diagnos_dictionary(dictionary_dataframe, file_path, key, file_name.lower())

                    cls.diagnos_mapping(joined_df, file_path, key, field_name, dictionary_dataframe, file_name.lower())

                    mapping_dicionaries_list.append(key)     

            spark.stop()

        except Exception as e:

            spark.stop()

            cls.print_failure_message(
                                    folder  = folder_name,
                                    partner = partner.lower(),
                                    job     = 'mmapping_gap' ,
                                    text = str(e)
                                    )

###################################################################################################################################
# This function will upload a pyspark df to mssql table in a specified server/dabase
###################################################################################################################################

    @classmethod


    def read_from_snowflake(cls, db_server,db_name, schema_name, table_name ):



        SNOWFLAKE_SOURCE_NAME = "net.snowflake.spark.snowflake"

        spark = SparkSession.builder \
                        .master("spark://master:7077") \
                        .appName("Upload") \
                        .getOrCreate()  



        pkb = cls.get_pkb()

        sf_options = {
            "sfURL": f"{db_server}.snowflakecomputing.com",
            "sfDatabase": db_name,
            "sfSchema": schema_name,
            "sfStage": SF_STAGE_NAME,
            "sfRole": SF_ROLE,
            "sfUser": SF_DB_USER,
            "pem_private_key": pkb  # <== path to your .p8 private key
        }
        
        

        get_tables_names_query = f"""
        SELECT *
        FROM {table_name}
        """




        table_df = spark.read \
            .format(SNOWFLAKE_SOURCE_NAME) \
            .options(**sf_options) \
            .option("query", get_tables_names_query) \
            .load()
        
        return table_df
###################################################################################################################################
# This function 
###################################################################################################################################

    @classmethod
    def get_pkb(cls):


            with open(SF_PRIVATE_KEY_PASSPHRASE_PATH, "r") as f:
                sf_private_key_passphrase = f.readline().strip()


            with open(SF_PRIVATE_KEY_PATH, "rb") as key:
                private_key = serialization.load_pem_private_key(
                    key.read(),
                    password=sf_private_key_passphrase.encode() if sf_private_key_passphrase else None,
                    backend=default_backend()
                )
        

            pkb = private_key.private_bytes(
                encoding=serialization.Encoding.PEM,
                format=serialization.PrivateFormat.PKCS8,
                encryption_algorithm=serialization.NoEncryption()
            )
            pkb = pkb.decode("UTF-8")
            pkb = re.sub("-*(BEGIN|END) PRIVATE KEY-*\n", "", pkb).replace("\n", "")

            return pkb

###################################################################################################################################
# This function will output the mapping gap for the passed formatted file name
###################################################################################################################################

   

    @classmethod
    def get_mothly_counts(cls,partner,file_name,table_name, file_path,folder_name):

        

        complete_file_path = file_path+'/formatted_'+file_name.lower()+'.csv'
        
        master_url = os.environ["SPARK_URL"].strip().replace('"', '')
        spark = SparkSession.builder.master(master_url).appName("Ovid").getOrCreate()
        # spark = SparkSession.builder.master("spark://spark-master:7077").appName("MONTHLY COUNT").getOrCreate()

        monthly_counts_out_file_path = file_path.replace('formatter','monthly_count')+table_name+'/'


        df = cls.spark_read(complete_file_path, spark)

        for col_name in df.columns:

            if 'DATE' in col_name:

                monthly_counts = df.groupBy(date_format(col_name, "yyyy-MM").alias("year_month")) \
                   .agg(count("*").alias("count")) \
                   .orderBy("year_month",  ascending=False)
                
                output_file_name = col_name.lower()+'_mothly_count.csv'

                cls.write_pyspark_output_file(monthly_counts,output_file_name, monthly_counts_out_file_path )




 ###################################################################################################################################
# This function will output the mapping gap for the passed formatted file name
###################################################################################################################################

   

    @classmethod
    def get_distinct_values(cls,partner,file_name,table_name, file_path,folder_name):

        bypass_files_list = ['fact_relationship']
        if file_name.lower() not in bypass_files_list:

            complete_file_path = file_path+'/formatted_'+file_name.lower()+'.csv'
            
            master_url = os.environ["SPARK_URL"].strip().replace('"', '')
            spark = SparkSession.builder.master(master_url).appName("Ovid").getOrCreate()


            # spark = SparkSession.builder.master("spark://spark-master:7077").appName("MONTHLY COUNT").getOrCreate()
            spark.conf.set("spark.sql.shuffle.partitions", 2000)  # or higher


            distinct_values_out_file_path = file_path.replace('formatter','distinct_values')+table_name+'/'


            df = cls.spark_read(complete_file_path, spark)

            total_count = df.count()

            for col_name in df.columns:

                if not col_name.endswith("ID") :

                    distinct_values_count = df.groupBy(col_name) \
                    .agg(count("*").alias("Count")) \
                    .withColumn("Percent", F.round((F.col("Count") / total_count * 100), 2))\
                    .orderBy("Count",  ascending=False)
                    
                    output_file_name = col_name.lower()+'_distinct_values.csv'

                    cls.write_pyspark_output_file(distinct_values_count,output_file_name, distinct_values_out_file_path )



           
###################################################################################################################################
# This function will output the mapping gap for the passed formatted file name
###################################################################################################################################

   

    @classmethod
    def get_mothly_counts(cls,partner,file_name,table_name, file_path,folder_name):

        

        complete_file_path = file_path+'/formatted_'+file_name.lower()+'.csv'
        
        master_url = os.environ["SPARK_URL"].strip().replace('"', '')
        spark = SparkSession.builder.master(master_url).appName("Ovid").getOrCreate()
        # spark = SparkSession.builder.master("spark://spark-master:7077").appName("MONTHLY COUNT").getOrCreate()

        monthly_counts_out_file_path = file_path.replace('formatter','monthly_count')+table_name+'/'


        df = cls.spark_read(complete_file_path, spark)

        for col_name in df.columns:

            if 'DATE' in col_name:

                monthly_counts = df.groupBy(date_format(col_name, "yyyy-MM").alias("year_month")) \
                   .agg(count("*").alias("count")) \
                   .orderBy("year_month",  ascending=False)
                
                output_file_name = col_name.lower()+'_mothly_count.csv'

                cls.write_pyspark_output_file(monthly_counts,output_file_name, monthly_counts_out_file_path )

          


###################################################################################################################################
# This function will diagnos a given dictionary data frame
########################################################################################################

    @classmethod
    def diagnos_dictionary(cls,dictionary_dataframe, file_path, file_name, table_name):

        # getting values that mapped to null or some flavor of null like 'OT, 'NI', and "UN
        filtered_df = dictionary_dataframe.filter((col("dict_value").isin("NI", "OT", "UN")) | (col("dict_value").isNull()) | (col("dict_value") == ""))

        out_file_path = file_path.replace('formatter','mapping_gap')+"/"+table_name+"/mapping_to_flavor_of_null/"
        output_file_name = file_name+"_flavor_of_null.csv"


        cls.write_pyspark_output_file(filtered_df, output_file_name, out_file_path )
           


###################################################################################################################################
# This function will diagnos a given joind dictionary dataframe
########################################################################################################

    @classmethod
    def diagnos_mapping(cls,joined_df, file_path, file_name, field_name, dictionary_dataframe, table_name):


        try:
            # output all mapping
            all_values_df = joined_df.select(joined_df[field_name],joined_df['dict_value'],joined_df['count'],joined_df['percent'])
            all_mapping_out_file_path = file_path.replace('formatter','mapping_gap')+"/"+table_name+"/all_mapping/"
            all_mapping_output_file_name = file_name+"_all_mapping.csv"

            cls.write_pyspark_output_file(all_values_df, all_mapping_output_file_name, all_mapping_out_file_path )


            
            # output missing mapping
            cdm_value_set = PcornetCDM.get_cdm_valueset_df(field_name)
            cdm_value_set = [x.strip() for x in cdm_value_set] # remove extra spaces from the values list

            missing_values_df = all_values_df.filter(col("dict_value").isNull() )
            
            
            missing_values_df_and_not_part_of_cdm = missing_values_df.filter(~col(field_name).isin(cdm_value_set))
            
            missing_mapping_out_file_path = file_path.replace('formatter','mapping_gap')+"/"+table_name+"/missing_mapping/"
            missing_mapping_output_file_name = file_name+"_missing_mapping.csv"
            cls.write_pyspark_output_file(missing_values_df_and_not_part_of_cdm, missing_mapping_output_file_name, missing_mapping_out_file_path )

            # output suggested mapping

            
            to_be_added_field = missing_values_df_and_not_part_of_cdm.select(

                    missing_values_df_and_not_part_of_cdm[field_name].alias('dict_key'),
                    lit('').alias('dict_value'),


            )

            unioned_df= dictionary_dataframe.union(to_be_added_field)
            suggested_mapping_out_file_path = file_path.replace('formatter','mapping_gap')+"/"+table_name+"/suggested_mapping/"

            out_file_name = suggested_mapping_out_file_path+"suggested_"+file_name+".py"

            result = os.makedirs(os.path.dirname(out_file_name), exist_ok=True)

            formatted_text = unioned_df.rdd \
                .map(lambda row: f'       {cls.process_key(row["dict_key"],row["dict_value"],file_name, 20)}  :  "{row["dict_value"]}",') 
            

            formatted_text= formatted_text.collect()
            

            # Join the formatted strings with newline characters
            formatted_text = "\n".join(formatted_text)

            # Write the formatted text to a text file
        
            first_line = "\n\n"+file_name+ " = {\n\n"
            last_line = "   \n\n}"

            with open(out_file_name, "w+") as file:
                file.write(first_line)
                file.write(formatted_text)
                file.write(last_line)


        except Exception as e:

            

            cls.print_with_style(str(e), 'danger red')
            
            
    
###################################################################################################################################
# 
########################################################################################################

    @classmethod
    def process_row(cls,row, file):


            print(row)


            key = cls.append_spaces(row.dict_key, 20),

            text = row.dict_value + '\n'

            file.write(text)


                
               

###################################################################################################################################
# 
########################################################################################################

    @classmethod
    def process_key(cls, key, value,file_name, size):

            key = str(key)

            if 'unit' not in file_name and 'UNIT' not in file_name:

                key = key.upper()

            key = '"'+key+'"'


            if value == "" or value == None:

                key = '# '+key

            while len(key) < size:

                key = key + ' '

            return key





###################################################################################################################################
# This function return and aggreated dataframe for a given table
###################################################################################################################################



    @classmethod
    def get_aggregate_value_from_table(cls, df, row_count, field_name):
        aggregated_df = df.groupBy(field_name).agg(count("*").alias("count"))

        if row_count != 0:
            percent_col = round((col("count") / row_count) * 100, 2).alias("percent")
        else:
            percent_col = lit(0).alias("percent")

        return aggregated_df.select(col(field_name), col("count"), percent_col)





###################################################################################################################################
# This function will
###################################################################################################################################

    @classmethod
    def db_upload(cls,db, db_server,file_name,table_name, file_path,folder_name,db_type,upload_type,merge_db,merge_schema):

        if db_type.upper() == 'PG':
            cls.db_upload_postgress(db, db_server,file_name,table_name, file_path,folder_name)

        if db_type.upper() == 'MSSQL':
            cls.db_upload_mssql(db, db_server,file_name,table_name, file_path,folder_name)

        if db_type.upper() == 'SF':
            cls.db_upload_snowflake(db, db_server,file_name,table_name, file_path,folder_name,upload_type,merge_db,merge_schema)




###################################################################################################################################
# This function will
###################################################################################################################################

    @classmethod
    def db_upload_cdm_complementary_tables(cls,db, db_server,folder_name,db_type, mappers_type):

        if db_type.upper() == 'PG':
            cls.db_upload_postgress(db, db_server,schemas[0],file_name,table_name, file_path,folder_name)

        if db_type.upper() == 'MSSQL':
            cls.db_upload_mssql(db, db_server,schemas[0],file_name,table_name, file_path,folder_name)

        if db_type.upper() == 'SF':
            cls.db_upload_cdm_complementary_tables_snowflake(db, db_server,folder_name, mappers_type)



###################################################################################################################################
# This function will upload a pyspark df to mssql table in a specified server/dabase
###################################################################################################################################

    @classmethod


    def db_upload_cdm_complementary_tables_snowflake(cls, db, db_server, folder_name, mappers_type):

        if '/' in folder_name:
            folder_name = folder_name.split('/')[1]

        SNOWFLAKE_SOURCE_NAME = "net.snowflake.spark.snowflake"
        schema_name = f"OVID_{folder_name.upper()}"



        spark = cls.get_spark_session("Upload")


        pkb = cls.get_pkb()

        sf_options = {
            "sfURL": f"{db_server}.snowflakecomputing.com",
            "sfDatabase": db,
            "sfSchema": "INFORMATION_SCHEMA",
            "sfStage": SF_STAGE_NAME,
            "sfRole": SF_ROLE,
            "sfUser": SF_DB_USER,
            "pem_private_key": pkb  # <== path to your .p8 private key
        }
        
        

        get_tables_names_query = f"""
        SELECT TABLE_NAME
        FROM TABLES
        WHERE TABLE_SCHEMA = '{schema_name}'
        """




        existing_tables_df = spark.read \
            .format(SNOWFLAKE_SOURCE_NAME) \
            .options(**sf_options) \
            .option("query", get_tables_names_query) \
            .load()


        tables_specs_path = f'/app/mapping_scripts/{mappers_type}/tables_specs.csv'

        cdm_df = spark.read.load(tables_specs_path,format="csv", sep=",", inferSchema="false", header="true", quote= '"')

        cdm_tables_names_df =  cdm_df.select("TABLE_NAME").distinct()


        missing_tables_names_df = cdm_tables_names_df.join(existing_tables_df, existing_tables_df["TABLE_NAME"] == cdm_tables_names_df["TABLE_NAME"], "left_anti")



        # Collect the table names into a Python list
        missing_table_names = [row["TABLE_NAME"] for row in missing_tables_names_df.collect()]

        # Iterate over the names
        for table in missing_table_names:
            
            schema = cls.get_snowflake_schema(table, spark)
            empty_df = spark.createDataFrame([], schema)

            sf_options = {
                "sfURL": f"{db_server}.snowflakecomputing.com",
                "sfDatabase": db,
                "sfSchema": f"OVID_{folder_name}",
                "sfStage": SF_STAGE_NAME,
                "sfRole": SF_ROLE,
                "sfUser": SF_DB_USER,
                "pem_private_key": pkb  # <== path to your .p8 private key
            }

            # Write DataFrame to Snowflake
            empty_df.write \
                .format(SNOWFLAKE_SOURCE_NAME) \
                .options(**sf_options) \
                .option("dbtable", table.upper()) \
                .option("createTableColumnTypes", schema) \
                .option("useSchema", f"OVID_{folder_name}") \
                .mode("ignore") \
                .save()
            

                    

###################################################################################################################################
# This function will upload a pyspark df to mssql table in a specified server/dabase
###################################################################################################################################

    @classmethod
    def merge_dfs(cls, merge_from_df,merge_into_df,table_name ):

        table_primary_key_array =  PCORNET_PRIMARY_KEYS.get(table_name)

        join_condition = [
                    merge_into_df[k] == merge_from_df[k]
                    for k in table_primary_key_array
                ]

        missing_rows = (
            merge_from_df.alias("from")
            .join(
                merge_into_df.alias("into"),
                on=join_condition,
                how="left_anti"
                )
            )
        

        merged_df = merge_into_df.unionByName(missing_rows)


        return merged_df



###################################################################################################################################
# This function will upload a pyspark df to mssql table in a specified server/dabase
###################################################################################################################################

    @classmethod

    def get_sf_schema(cls, df, table_name):





        """
        Generate Snowflake schema from a Spark DataFrame based on PCORnet CDM.
        
        Args:
            df (DataFrame): Spark DataFrame containing the table data.
            table_name (str): Name of the table to get CDM info.
        
        Returns:
            List[StructField]: Snowflake schema fields.
        """
        schema_sf = []
        cdm_fields = []

        try:
            spark = cls.get_spark_session("Get Schema")

            # Get CDM fields for the table
            cdm_df = PcornetCDM.get_cdm_df(spark)
            table_cdm_df = cdm_df.filter(cdm_df["TABLE_NAME"] == table_name)
            table_cdm = table_cdm_df.collect()

            # Add CDM-defined fields
            for row in table_cdm:
                schema_sf.append(cls.get_df_field_config(df, row["FIELD_NAME"], row["RDBMS_DATA_TYPE"]))
                cdm_fields.append(row["FIELD_NAME"])

            # Add default fields

            # Add any extra fields from the DataFrame that are not in CDM
            for field in df.columns:
                if field not in cdm_fields and field.upper() not in ['SOURCE', 'UPDATED']:
                    field_size = int(cls.get_max_feild_size_from_df(df, field))
                    schema_sf.append(StructField(field, StringType(), metadata={"maxlength": field_size}))

            schema_sf.append(StructField("UPDATED", StringType(), metadata={"maxlength": 19}))
            schema_sf.append(StructField("SOURCE", StringType(), metadata={"maxlength": 20}))

            # spark.stop()

        except Exception as e:
            # spark.stop()
            cls.print_failure_message(
                folder='',
                partner='',
                job='',
                text=str(e)
            )

        return schema_sf





###################################################################################################################################
# This function will upload a pyspark df to mssql table in a specified server/dabase
###################################################################################################################################

    @classmethod


    def db_upload_snowflake(cls, db, db_server, file_name, table_name, file_path, folder_name, upload_type,merge_db,merge_schema):

        if '/' in folder_name:
            folder_name = folder_name.split('/')[1]

        SNOWFLAKE_SOURCE_NAME = "net.snowflake.spark.snowflake"
        
        try:
            
            spark = cls.get_spark_session("Upload")
            spark.conf.set("spark.sql.legacy.timeParserPolicy", "LEGACY")

            # custom_schema = StructType(schema)

            # file_df = spark.read.load(file_path+file_name+'*', format="csv", sep="\t", header="true", quote='"')

            file_df = cls.spark_read(file_path+file_name+'*', spark)






            pkb = cls.get_pkb()

            
            if upload_type == 'merge' and table_name.upper()!= 'HARVEST' :

                db_schema_name = f"OVID_{folder_name}_merged_with_{merge_db}_{merge_schema}"

                sf_options = {
                "sfURL": f"{db_server}.snowflakecomputing.com",
                "sfDatabase": db,
                "sfSchema": db_schema_name,
                "sfStage": SF_STAGE_NAME,
                "sfRole": SF_ROLE,
                "sfUser": SF_DB_USER,
                "pem_private_key": pkb  # <== path to your .p8 private key
            }


                merge_df = (
                        spark.read
                            .format(SNOWFLAKE_SOURCE_NAME)
                            .options(**sf_options)
                            .option("dbtable", f"{merge_db}.{merge_schema}.{table_name}")
                            .load()
                    )



                # merge_df.show()

                file_cols = file_df.columns

                merge_df = merge_df.select(
                    *[col(c) for c in file_cols if c in merge_df.columns]
                )


                # merge_df.show()


                merged_df = cls.merge_dfs(
                    
                    merge_from_df = merge_df,
                    merge_into_df = file_df,
                    table_name    = table_name )


                schema_type = cls.get_sf_schema(merged_df, table_name)
                custom_schema = StructType(schema_type)

                conn = cls.get_sf_connection(db_server)
                cur = conn.cursor()
                pkb = cls.get_pkb()

                cur.execute(f"CREATE SCHEMA IF NOT EXISTS {db}.{db_schema_name}")


                for field in schema_type:
                    if isinstance(field.dataType, DateType) and field.name in merged_df.columns:
                        merged_df = merged_df.withColumn(field.name, to_date(col(field.name), "yyyy-MM-dd"))

                    if isinstance(field.dataType, DoubleType) and field.name in merged_df.columns:
                        merged_df = merged_df.withColumn(field.name, col(field.name).cast(DoubleType()))

                merged_df_with_schema = spark.createDataFrame(merged_df.rdd, custom_schema)

                # Write DataFrame to Snowflake
                merged_df_with_schema.repartition(10).write \
                    .format(SNOWFLAKE_SOURCE_NAME) \
                    .options(**sf_options) \
                    .option("on_error", "CONTINUE") \
                    .option("dbtable", table_name.upper()) \
                    .option("createTableColumnTypes", schema_type) \
                    .option("useSchema", db_schema_name) \
                    .mode("overwrite") \
                    .save()



            else:


                schema_type = cls.get_sf_schema(file_df,table_name)

                custom_schema = StructType(schema_type)


                db_schema_name = f'OVID_{folder_name}'
                

                conn = cls.get_sf_connection(db_server)
                cur = conn.cursor()
                pkb = cls.get_pkb()
                cur.execute(f"CREATE SCHEMA IF NOT EXISTS {db}.{db_schema_name}")

                for field in schema_type:
                    if isinstance(field.dataType, DateType) and field.name in file_df.columns:
                        file_df = file_df.withColumn(field.name, to_date(col(field.name), "yyyy-MM-dd"))

                    if isinstance(field.dataType, DoubleType) and field.name in file_df.columns:
                        file_df = file_df.withColumn(field.name, col(field.name).cast(DoubleType()))

                file_df_with_schema = spark.createDataFrame(file_df.rdd, custom_schema)

                sf_options = {
                "sfURL": f"{db_server}.snowflakecomputing.com",
                "sfDatabase": db,
                "sfSchema": db_schema_name,
                "sfStage": SF_STAGE_NAME,
                "sfRole": SF_ROLE,
                "sfUser": SF_DB_USER,
                "pem_private_key": pkb  # <== path to your .p8 private key
            }

                # Write DataFrame to Snowflake
                file_df_with_schema.repartition(10).write \
                    .format(SNOWFLAKE_SOURCE_NAME) \
                    .options(**sf_options) \
                    .option("on_error", "CONTINUE") \
                    .option("dbtable", table_name.upper()) \
                    .option("createTableColumnTypes", schema_type) \
                    .option("useSchema", db_schema_name) \
                    .mode("overwrite") \
                    .save()
                
            spark.stop()
            cur.close()
            conn.close()
        except Exception as e:

            spark.stop()            
            cls.print_with_style(str(e), 'danger red')




###################################################################################################################################
# This function 
###################################################################################################################################

    @classmethod
    def get_pkb(cls):

            SF_PRIVATE_KEY_PASSPHRASE_PATH =  "/ovid_credentials/" + SF_PRIVATE_KEY_PASSPHRASE_FILE_NAME

            with open(SF_PRIVATE_KEY_PASSPHRASE_PATH, "r") as f:
                sf_private_key_passphrase = f.readline().strip()

            SF_PRIVATE_KEY_PATH =  "/ovid_credentials/" + SF_PRIVATE_KEY_FILE_NAME

            with open(SF_PRIVATE_KEY_PATH, "rb") as key:
                private_key = serialization.load_pem_private_key(
                    key.read(),
                    password=sf_private_key_passphrase.encode() if sf_private_key_passphrase else None,
                    backend=default_backend()
                )
        

            pkb = private_key.private_bytes(
                encoding=serialization.Encoding.PEM,
                format=serialization.PrivateFormat.PKCS8,
                encryption_algorithm=serialization.NoEncryption()
            )
            pkb = pkb.decode("UTF-8")
            pkb = re.sub("-*(BEGIN|END) PRIVATE KEY-*\n", "", pkb).replace("\n", "")

            return pkb


###################################################################################################################################
# This function 
###################################################################################################################################

    @classmethod
    def create_scheduler_db(cls):
       
        
        try:
         
            conn = cls.get_sf_connection(SCHEDULER_DB_SERVER)
            cur = conn.cursor()
            cur.execute(f"CREATE DATABASE {SCHEDULER_DB_NAME}")
            cur.execute(f"CREATE SCHEMA {SCHEDULER_DB_NAME}.{SCHEDULER_DB_SCHEMA}")
            cur.execute(f"{CREATE_BATCH_TABLE_DLL}")
            cur.execute(f"{CREATE_BATCH_TABLE_FILE_DLL}")

            conn.close()

              
                
        except Exception as e:

            cls.print_with_style(str(e), 'danger red')




###################################################################################################################################
# This function 
###################################################################################################################################

    @classmethod
    def execute_query(cls, query):
       
        
        try:
        
            conn = cls.get_sf_connection(SCHEDULER_DB_SERVER)
            cur = conn.cursor()
            result = cur.execute(query).fetchall()
   
            cur.close()
            conn.close()

            return result            
        except Exception as e:

            cls.print_with_style(str(e), 'danger red')


###################################################################################################################################
# This function 
###################################################################################################################################

    @classmethod
    def get_sf_connection(cls, DB_SERVER):

        SNOWFLAKE_SOURCE_NAME = "net.snowflake.spark.snowflake"
        
        SF_PRIVATE_KEY_PASSPHRASE_PATH =  "/ovid_credentials/" + SF_PRIVATE_KEY_PASSPHRASE_FILE_NAME

        with open(SF_PRIVATE_KEY_PASSPHRASE_PATH, "r") as f:
                sf_private_key_passphrase = f.readline().strip()

        SF_PRIVATE_KEY_PATH =  "/ovid_credentials/" + SF_PRIVATE_KEY_FILE_NAME

        with open(SF_PRIVATE_KEY_PATH, "rb") as key:
                private_key = serialization.load_pem_private_key(
                    key.read(),
                    password=sf_private_key_passphrase.encode() if sf_private_key_passphrase else None,
                    backend=default_backend()
                )
            
            # Convert to DER format for Snowflake
        private_key_bytes = private_key.private_bytes(
                encoding=serialization.Encoding.DER,
                format=serialization.PrivateFormat.PKCS8,
                encryption_algorithm=serialization.NoEncryption()
            )
            
            # Connect to Snowflake using private key
        conn = snowflake.connector.connect(
                user=SF_DB_USER,
                private_key=private_key_bytes,
                account=DB_SERVER
            )
        
        return conn

###################################################################################################################################
# This function 
###################################################################################################################################

    @classmethod
    def check_scheduler_db_exists(cls):

        try:
            conn = cls.get_sf_connection(SCHEDULER_DB_SERVER)
            cursor = conn.cursor()
            try:
                query = f"SHOW DATABASES LIKE '{SCHEDULER_DB_NAME.upper()}'"
                cursor.execute(query)  # Snowflake stores names in uppercase
                result = cursor.fetchone()  # Fetch once and store result
            finally:
                cursor.close()
                conn.close()

            return result is not None
        
        except Exception as e:
            print(f"Error checking database existence: {e}")
            return False



###################################################################################################################################
# This function will upload a pyspark df to mssql table in a specified server/dabase
###################################################################################################################################

    @classmethod
    def db_upload_mssql(cls,db, db_server,schema,file_name,table_name, file_path,folder_name):


        try:

            master_url = os.environ["SPARK_URL"].strip().replace('"', '')
            spark = SparkSession.builder.master(master_url).appName("Ovid").getOrCreate()
            # spark = SparkSession.builder.master("spark://spark-master:7077").appName("MSSQL Upload").getOrCreate()

            conf = SparkConf()
            conf.set("spark.driver.extraClassPath", "ovid_setup/jars/mssql-jdbc-driver.jar")
            file_df = spark.read.option("inferSchema", "false").load(file_path+file_name+'*',format="csv", sep="\t", inferSchema="false", header="true",  quote= '"')
            
            string_columns = [f'CAST(`{col}` AS STRING) AS `{col}`' for col in file_df.columns]
            result_df = file_df.selectExpr(*string_columns)

            
            jdbc_url = f"jdbc:sqlserver://{db_server};databaseName={db}"
            
            if schema != None:
                properties = {
                    "user": MSSQL_DB_USER,
                    "password": MSSQL_DB_PASS,
                    "driver": "com.microsoft.sqlserver.jdbc.SQLServerDriver",
                    "trustServerCertificate": "true",
                    "createTableColumnTypes": schema}
            else:
                properties = {
                    "user": MSSQL_DB_USER,
                    "password": MSSQL_DB_PASS,
                    "driver": "com.microsoft.sqlserver.jdbc.SQLServerDriver",
                    "trustServerCertificate": "true"}


            table_name = f"OVID_{folder_name}_{table_name}"
            
            result_df.write.jdbc(url=jdbc_url, table=table_name, mode="overwrite", properties=properties)

            spark.stop()
        except Exception as e:

            spark.stop()
            

            cls.print_with_style(str(e), 'danger red')




###################################################################################################################################
# This function will upload a pyspark df to postgress table in a specified server/dabase
###################################################################################################################################

    @classmethod
    def db_upload_postgress(cls,db, db_server,schema,file_name,table_name, file_path,folder_name):
                #Boot Cluster
        start = time.time()
        date_format_pattern = "yyyy-MM-dd"
        master_url = os.environ["SPARK_URL"].strip().replace('"', '')
        spark = SparkSession.builder.master(master_url).appName("Ovid").getOrCreate()
        # spark = SparkSession.builder.master("spark://spark-master:7077").appName("PSQL Upload").getOrCreate()

        try:
            conf = SparkConf()
            conf.set("spark.driver.extraClassPath", "ovid_setup/jars/postgresql-42.7.4.jar")

            # Parse schema to get column names and types
            column_types = {}
            for line in schema.strip().split(",\n"):
                match = re.match(r"(\w+)\s+(\w+)", line.strip())
                if match:
                    column, dtype = match.groups()
                    column_types[column] = dtype

            # Load your file into a DataFrame
            file_df = spark.read.option("inferSchema", "false").csv(file_path + file_name + '*', sep="\t", header=True, quote='"')


            # Create casting expressions based on parsed schema
            casting_expressions = [
                f'CAST(`{col}` AS {"STRING" if dtype.startswith("VARCHAR") else dtype}) AS `{col}`'
                for col, dtype in column_types.items()
            ]

            # Apply casting expressions
            result_df = file_df.selectExpr(*casting_expressions)
            
            if schema != None:
                properties = {
                    "user": PG_DB_USER,
                    "password": PG_DB_PASS,
                    "driver": "org.postgresql.Driver",
                    "stringtype":"unspecified",
                    "createTableColumnTypes": schema}
            else:
                properties = {
                    "user": PG_DB_USER,
                    "password": PG_DB_PASS,
                    "driver": "org.postgresql.Driver",
                   }


            table_name = f"OVID_{folder_name}_{table_name}"
            jdbc_url = f"jdbc:postgresql://{db_server}/{db}"
            result_df.write.jdbc(url=jdbc_url, table=table_name, mode="overwrite", properties=properties)

            spark.stop()
        except Exception as e:

             spark.stop()
            

             cls.print_with_style(str(e), 'danger red')



###################################################################################################################################
# This function will return a list of folder in a given path
###################################################################################################################################
    @classmethod
    def get_folders_list(cls, path):

       
        
        folders_names = os.listdir(path)

        folders_names_list = [file for file in folders_names if file]

        return folders_names_list
###################################################################################################################################
# This 
###################################################################################################################################
    @classmethod
    def clean_excel_value(cls,val):

        if isinstance(val, str):

            # Remove illegal Unicode control characters (0x00 to 0x1F except newline/tab)
            val = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", val)

        return val
###################################################################################################################################
# This function generate a mapping report as a form of an excel file
###################################################################################################################################


    @classmethod
    def generate_mapping_report(cls,mapping_gap_output_folder_path, folder, partner):


        try:

     
            if '/' in folder :
                folder = folder.split('/')[1]


            spark = cls.get_spark_session('generate mapping report')

            empty_columns = pd.DataFrame(columns=[""] * 2)
            # Path to the Excel file
            excel_file_path = mapping_gap_output_folder_path + partner+'_'+folder+"_mapping_report.xlsx"


            # Initialize a workbook object
            wb = openpyxl.Workbook()



            tables_list = cls.get_folders_list(mapping_gap_output_folder_path)

            for table in tables_list :

            


                if 'mapping_report' not in table:

                    safe_sheet_name = re.sub(r'[:\\/*?[\]#]', '_', table)[:31]  # Excel allows max 31 characters

                    ws = wb.create_sheet(title=safe_sheet_name)

                    combined_pd = empty_columns


                    table_all_mappings_path = mapping_gap_output_folder_path +"/"+safe_sheet_name+"/all_mapping/"

                    table_all_mapping_files_list = cls.get_folders_list(table_all_mappings_path)

                    for mapping_file in table_all_mapping_files_list:

                        mapping_df =  cls.spark_read(table_all_mappings_path+mapping_file, spark)

                        mapping_df = mapping_df.limit(1000000)


                        pandas_df = mapping_df.toPandas()
                        combined_pd = pd.concat([combined_pd, empty_columns, pandas_df], axis=1)


                        for col_idx, col_name in enumerate(combined_pd.columns, 1):
                            ws.cell(row=1, column=col_idx, value=col_name)


                        for row_idx, row in enumerate(combined_pd.values, 2):  # Start from row 2 (after header row)
                            for col_idx, value in enumerate(row, 1):

                                value = cls.clean_excel_value(value)

                                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                                # Calculate the width of the cell based on the length of the value
                                if isinstance(value, str):
                                    width = len(value)
                                else:
                                    width = len(str(value))

                                # Update the column width if necessary
                                if width > ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width:
                                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width


            if 'Sheet' in wb.sheetnames:
                # Remove the sheet
                wb.remove(wb['Sheet'])

            wb.save(excel_file_path)




            spark.stop()


        except Exception as e:

            spark.stop()
            cls.print_failure_message(
                                    folder  = folder,
                                    partner = partner.lower(),
                                    job     = 'mapping_report' ,
                                    text = str(e)
                                    )




###################################################################################################################################
# This function will return a pysark dataframe for a givin file path
###################################################################################################################################


    @classmethod
    def spark_read(cls,file_path, spark):

        # return spark.read.option("inferSchema", "false").load(file_path,format="csv",   sep="\t", inferSchema="false", header="true",  quote= '"')


        return spark.read.csv(
                file_path,
                sep="\t",
                header=True,
                quote='"',
                inferSchema=False,
                multiLine=False,
                escape="\""
            )




###################################################################################################################################
# This function will append any additional fields that are not in the Pcornet CDM to the mapper output
###################################################################################################################################

   

    @classmethod
    def append_additional_fields(cls,mapped_df, file_name, deduplicated_data_folder_path,join_field, spark ):


        # read the deduplicated file
        deduplicated_file = spark.read.load(deduplicated_data_folder_path+file_name,format="csv", sep="\t", inferSchema="true", header="true", quote= '"')
        
        if len(deduplicated_file.columns) + 3 > len(mapped_df.columns) :

            #get the headers from the deduplicated and the mapped files for comparison
            deduplicated_file_columns = deduplicated_file.columns
            mapped_df_columns = mapped_df.columns


            # Identify common headers
            common_headers = set(mapped_df_columns).intersection(deduplicated_file_columns)
            common_headers_without_the_join_field = [item for item in common_headers if item != join_field]

            # Drop common headers from the second DataFrame
            additional_fields_data = deduplicated_file.drop(*common_headers_without_the_join_field)
            
            #Renaming the join field to avoid conflect during the join
            join_field_temp_name = join_field+"_new"

            additional_fields_data = additional_fields_data.withColumnRenamed(join_field, join_field_temp_name)
            
            # Join DataFrames based on the "Name" column
            appended_payspark_data_frame = mapped_df.join(additional_fields_data, mapped_df["JOIN_FIELD"]==additional_fields_data[join_field_temp_name])
        
        #Dropping the temporary join fields

            columns_to_drop = ["JOIN_FIELD",join_field_temp_name ]
            appended_payspark_data_frame = appended_payspark_data_frame.drop(*columns_to_drop)

            return appended_payspark_data_frame
        
        else:

            return mapped_df.drop("JOIN_FIELD")
        


###################################################################################################################################
# This function will convert all known time formats to the '%H:%M'format
###################################################################################################################################

    @staticmethod
    def format_time( val):
        """ Convenience method to try all known formats """

        try:
            parsed_time = dp.parse(val)
        except Exception as exc:
           
            return ""
        return parsed_time.strftime('%H:%M')


###################################################################################################################################
# This will ensure the value is a float or it will return None
###################################################################################################################################

    @staticmethod
    def copy_if_float( val):
        """ Convenience method to try all known formats """
        if not val:
            return None
        try:
            return float(val)
        except:

            return None



###################################################################################################################################
# This function will convert all known date formats to the '%Y-%m-%d format
###################################################################################################################################

   
    @staticmethod
    def format_date( val):
        """ Convenience method to try all known formats """
        print(val)
        try:
            parsed_date = dp.parse(val)
        except:

            print(val)
           
            return None

        return parsed_date.strftime('%Y-%m-%d')




###################################################################################################################################
# This function will return the time portion from a datetime value
###################################################################################################################################

   
    @staticmethod
    def get_time_from_datetime(  dateTime):
        
            try:
                
                return str(dateTime.time().strftime('%H:%M:%S'))[0:5]
            except:
                return None

###################################################################################################################################
# This function will return the date portion from a datetime value
###################################################################################################################################

   
    @staticmethod
    def get_date_from_datetime(  dateTime):
        
        try:
            
            return str(dateTime.strftime("%Y-%m-%d"))
        except:
            return None

        
  
###################################################################################################################################
# This function will validate if a giving partenr name is valid or not
###################################################################################################################################

   

    @classmethod
    def valid_partner_name(cls, input_partner_name ):

        
        partners_root = "/app/partners"

        if not os.path.exists(partners_root):
            return None

        partners_list = [
            name for name in os.listdir(partners_root)
            if os.path.isdir(os.path.join(partners_root, name))
        ]

        if input_partner_name in partners_list:
            return True
        else:
            return None
###################################################################################################################################
# This function will create and return a spark session
###################################################################################################################################

    @classmethod
    def get_spark_session(cls, appName):
        try:
            spark.stop()
        except:
            None

        spark = SparkSession.builder.master(os.environ["SPARK_URL"]).appName(appName).config("spark.sql.parquet.datetimeRebaseModeInRead", "LEGACY").getOrCreate()
        spark.sparkContext.setLogLevel('OFF')

        return spark

###################################################################################################################################
    @classmethod
    def print_failure_message(cls, folder,partner, job, text):


        current_utc_datetime = datetime.utcnow()
        new_york_tz = pytz.timezone('America/New_York')

        current_ny_datetime = current_utc_datetime.replace(tzinfo=pytz.utc).astimezone(new_york_tz)

        formatted_ny_datetime = current_ny_datetime.strftime("%Y-%m-%d %H:%M:%S %Z")

        message = formatted_ny_datetime


        while len(message) < 39:
            message = message + ' '


        message = message + "--Partner: " + partner

    
        while len(message) < 55:
            message = message + ' '


        message = message + "--Folder: " + folder

    
        while len(message) < 80:
            message = message + ' '
        

        message = message+' --Running: '+ job+ ' Failed!!!' 

        while len(message) < 125:
            message = message + ' '

        while len(message) < 145:
            message = message + '.'

        if "Path does not exist" in text:

            color = 'warning orange'

        else:

            color = 'danger red'

        cls.print_with_style(message, color)
        cls.print_with_style(text, color)




###################################################################################################################################
# This 
###################################################################################################################################
    @classmethod
    def get_max_feild_size_from_df(cls, df,field_name):



        max_length = str(df.select(max(length(df[field_name]))).collect()[0][0])


        try:
            if int(max_length) >= 1:
               pass
        except:
            max_length = "1"


        return max_length





###################################################################################################################################
# This 
###################################################################################################################################
    @classmethod
    def get_field_config(cls, file_path, field_name, field_data_type, spark):

        harmonized_size =0 

        field_max_size = int(cls.get_max_feild_size_from_df(file_path,field_name, spark))


        if 'RDBMS Text' in field_data_type:

                field_type = StringType()

                cdm_field_size_list = field_data_type.split('(')

                cdm_field_size = cdm_field_size_list[len(cdm_field_size_list)-1].replace(')','')

                if 'x' in cdm_field_size:
                    
                    if 'RAW' in field_name:


                        cdm_field_size  = cls.get_max_feild_size_from_df(file_path,field_name, spark)


                    else:
                    

                        cdm_field_size = harmonized_filed_sizes_dict.get(field_name,0)
                
                

                if int(cdm_field_size) > int(field_max_size):
                        field_size = cdm_field_size
                else:
                        field_size = field_max_size





                field_config = StructField(field_name, field_type, metadata={"maxlength":int(field_size)})


                    

        if 'RDBMS Date' in field_data_type:

            field_type = DateType()

            field_config =   StructField(field_name, field_type, True)   

        if 'RDBMS Number' in field_data_type:

            field_type = DoubleType() 

            field_config =    StructField(field_name, field_type, True)   


        return field_config




###################################################################################################################################
# This 
###################################################################################################################################


    @classmethod
    def get_df_field_config(cls, df, field_name, field_data_type):
        """
        Generate a StructField for Snowflake schema based on CDM field info and actual DataFrame.

        Args:
            df (DataFrame): Spark DataFrame containing the table data.
            field_name (str): Field/column name.
            field_data_type (str): CDM RDBMS data type (e.g., 'RDBMS Text(50)').
            spark (SparkSession): Spark session.

        Returns:
            StructField: Field configuration for Snowflake schema.
        """
        harmonized_size = 0

        # Get max size from the DataFrame
        field_max_size = int(cls.get_max_feild_size_from_df(df, field_name))

        if 'RDBMS Text' in field_data_type:
            field_type = StringType()

            # Extract CDM field size from string like "RDBMS Text(50)"
            cdm_field_size_list = field_data_type.split('(')
            cdm_field_size = cdm_field_size_list[-1].replace(')', '')

            # Handle "x" notation or RAW fields
            if 'x' in cdm_field_size:
                if 'RAW' in field_name.upper():
                    cdm_field_size = cls.get_max_feild_size_from_df(df, field_name)
                else:
                    cdm_field_size = harmonized_filed_sizes_dict.get(field_name, 0)

            # Pick the larger of CDM size vs actual data size
            # field_size = max(int(cdm_field_size), int(field_max_size))

            if int(cdm_field_size) > int(field_max_size):

                field_size = int(cdm_field_size)
            else:

                field_size = int(field_max_size)

            field_config = StructField(field_name, field_type, metadata={"maxlength": int(field_size)})

        elif 'RDBMS Date' in field_data_type:
            field_type = DateType()
            field_config = StructField(field_name, field_type, True)

        elif 'RDBMS Number' in field_data_type:
            field_type = DoubleType()
            field_config = StructField(field_name, field_type, True)

        else:
            # Default to string if type is unknown
            field_config = StructField(field_name, StringType(), True)

        return field_config


###################################################################################################################################
# This 
###################################################################################################################################
    @classmethod
    def get_field_config_snowflake(cls, field_name, field_data_type, spark):

  



        if 'RDBMS Text' in field_data_type:

                field_type = StringType()

                cdm_field_size_list = field_data_type.split('(')

                cdm_field_size = cdm_field_size_list[len(cdm_field_size_list)-1].replace(')','')

                if 'x' in cdm_field_size:
                    
                    if 'RAW' in field_name:


                        cdm_field_size  = 20


                    else:
                    

                        cdm_field_size = harmonized_filed_sizes_dict.get(field_name,20)
                
                





                field_config = StructField(field_name, field_type, metadata={"maxlength":int(cdm_field_size)})


                    

        if 'RDBMS Date' in field_data_type:

            field_type = DateType()

            field_config =   StructField(field_name, field_type, True)   

        if 'RDBMS Number' in field_data_type:

            field_type = DoubleType() 

            field_config =    StructField(field_name, field_type, True)   


        return field_config


###################################################################################################################################
# This 
###################################################################################################################################
    @classmethod
    def get_schema_line(cls, file_path, field_name, field_data_type, spark):

        harmonized_size =0 

        field_max_size = int(cls.get_max_feild_size_from_df(file_path,field_name, spark))


        if 'RDBMS Text' in field_data_type:

                schema_type = 'VARCHAR'

                cdm_field_size_list = field_data_type.split('(')

                cdm_field_size = cdm_field_size_list[len(cdm_field_size_list)-1].replace(')','')

                if 'x' in cdm_field_size:
                    
                    if 'RAW' in field_name:


                        cdm_field_size  = cls.get_max_feild_size_from_df(file_path,field_name, spark)

                    else:
                    

                        cdm_field_size = harmonized_filed_sizes_dict.get(field_name,0)
                


                if int(cdm_field_size) > int(field_max_size):
                        field_size = cdm_field_size
                else:
                        field_size = field_max_size



                schema_line =     field_name +' '+ schema_type+"("+str(field_size)+"),\n"


                    

        if 'RDBMS Date' in field_data_type:

            schema_type = 'DATE'

            schema_line =     field_name +' '+ schema_type+",\n"

        if 'RDBMS Number' in field_data_type:

            schema_type = 'FLOAT'

            schema_line =     field_name +' '+ schema_type+",\n"


        return schema_line

        


        


###################################################################################################################################
# This function will adjust the schema fields depend on the size specified by the user or the max size in the field
###################################################################################################################################
    @classmethod
    def get_schemas(cls, file_name, table_path):






        file_path = table_path+'fixed_'+file_name.lower()+".csv"

        schema_mssql = ""
        schema_sf    = []
        schemas      = []
        cdm_fields   = []

        try:

            spark = cls.get_spark_session("Get Schema")


            cdm_df = PcornetCDM.get_cdm_df(spark)

            table_cdm_df = cdm_df.filter(cdm_df["TABLE_NAME"]==file_name)

            table_cdm = table_cdm_df.collect()

            fixed_table_df = cls.spark_read(file_path,spark)
    



            for row in table_cdm:


                schema_mssql = schema_mssql + '\n' + cls.get_schema_line(file_path,row["FIELD_NAME"],row["RDBMS_DATA_TYPE"] , spark) 
                schema_sf.append( cls.get_field_config(file_path,row["FIELD_NAME"],row["RDBMS_DATA_TYPE"] , spark) )
                cdm_fields.append(row["FIELD_NAME"])



            

            schema_mssql = schema_mssql +"UPDATED VARCHAR(19),\n"
            schema_mssql = schema_mssql +"SOURCE VARCHAR(20)"
            schema_sf.append(StructField("UPDATED", StringType(), metadata={"maxlength":19}),)
            schema_sf.append(StructField("SOURCE", StringType(), metadata={"maxlength":20}),)

            

            for field in fixed_table_df.columns:

                if field not in cdm_fields and field.upper() not in ['SOURCE','UPDATED']:

                    field_size = int(cls.get_max_feild_size_from_df(file_path, field, spark))

                    schema_sf.append(StructField(field, StringType(), metadata={"maxlength":field_size}),)
                    schema_mssql = schema_mssql +F",\n{field} VARCHAR({field_size})"



            schemas =  [schema_mssql,schema_sf]

            spark.stop()

        except Exception as e:
           
            spark.stop() 
            cf.print_failure_message(
                                    folder  = '',
                                    partner = '',
                                    job     = '' ,
                                    text = str(e)
                                    )

        return schemas


###################################################################################################################################
# This function will get snowflake schema for a table
###################################################################################################################################
    @classmethod
    def get_snowflake_schema(cls, table_name, spark):

            schema_sf = []
            cdm_df = PcornetCDM.get_cdm_df(spark)

            table_cdm_df = cdm_df.filter(cdm_df["TABLE_NAME"]==table_name)

            # table_cdm_df.show() 



            for row in table_cdm_df.collect():


                schema_sf.append( cls.get_field_config_snowflake(row["FIELD_NAME"],row["RDBMS_DATA_TYPE"] , spark) )


            schema_sf.append(StructField("UPDATED", StringType(), metadata={"maxlength":100}),)
            schema_sf.append(StructField("SOURCE", StringType(), metadata={"maxlength":20}),)
            schema_sf = StructType(schema_sf)

            return schema_sf     




###################################################################################################################################
# This function will verify the passed value is a float or it will return null
###################################################################################################################################
    @classmethod

    def return_if_float(cls, val):
        try:
            float(val)
            return val
        except:
            #Not a float
            return None
    



###################################################################################################################################
# This function will get all the created mappers python script and return a list of thier names
###################################################################################################################################
    @classmethod
    def print_run_status(cls,job_num,total_jobs_count, job, folder, partner):
        


        current_utc_datetime = datetime.utcnow()
        new_york_tz = pytz.timezone('America/New_York')

        current_ny_datetime = current_utc_datetime.replace(tzinfo=pytz.utc).astimezone(new_york_tz)

        formatted_ny_datetime = current_ny_datetime.strftime("%Y-%m-%d %H:%M:%S %Z")

        message = formatted_ny_datetime

        while len(message) < 25:
            message = message + ' '

        message = message + 'Job: '+str(job_num)+'/'+str(total_jobs_count)
        while len(message) < 39:
            message = message + ' '

        message = message + "--Partner: " + partner

    
        while len(message) < 55:
            message = message + ' '


        message = message + "--Folder: " + folder

    
        while len(message) < 80:
            message = message + ' '
        

        message = message+' --Running: '+job 

        while len(message) < 125:
            message = message + ' '

        while len(message) < 145:
            message = message + '.'


        cls.print_with_style(message, 'matrix green')

###################################################################################################################################
# This function will print a text with matrix green color
###################################################################################################################################


    @classmethod
    def print_with_style(cls, text, color):

        if color == 'matrix green':

            bold_code = '\033[1m'
            color_code = '\033[92m'
            reset_code = '\033[0m'  # Reset to default style
       
        if color == 'danger red':

            bold_code = '\033[1m'
            color_code = RED='\033[0;31m'
            reset_code = '\033[0m'  # Reset to default style


        if color == 'warning orange':

            bold_code = '\033[1m'
            color_code = '\033[38;5;214m'  # Orange
            reset_code = '\033[0m'  # Reset to default style


        if color == 'blue':
            bold_code = '\033[1m'
            color_code = BLUE='\033[0;34m'
            reset_code = '\033[0m'  # Reset to default style

        if color == 'pomegranate':
            bold_code = '\033[1m'
            color_code =  POMEGRANATE='\033[0;91m'
            reset_code = '\033[0m'  # Reset to default style

        if color == 'dark pink':
            bold_code = '\033[1m'
            color_code =   DARK_PINK='\033[0;35m'
            reset_code = '\033[0m'  # Reset to default style
   
        # Print bold green text
        print(bold_code + color_code + text + reset_code)

###################################################################################################################################
# This function will print a text with matrix green color
###################################################################################################################################


    @classmethod
    def print_fixer_status(cls,current_count,total_count, fix_type, fix_name):
        


        current_utc_datetime = datetime.utcnow()
        new_york_tz = pytz.timezone('America/New_York')

        current_ny_datetime = current_utc_datetime.replace(tzinfo=pytz.utc).astimezone(new_york_tz)

        formatted_ny_datetime = current_ny_datetime.strftime("%Y-%m-%d %H:%M:%S %Z")

        message = formatted_ny_datetime

        while len(message) < 25:
            message = message + ' '

        message = message + 'Fix: '+str(current_count)+'/'+str(total_count)
        while len(message) < 39:
            message = message + ' '

        message = message + "--Fix type: " + fix_type+ ' '

    
        while len(message) < 80:
            message = message + ' '


        message = message + "--Fix Name: " +  fix_name+ ' '


        while len(message) < 125:
            message = message + ' '

        while len(message) < 145:
            message = message + '.'

        cls.print_with_style(message, 'dark pink')

##################################################################################################################################
# 
###################################################################################################################################

    @staticmethod
    def get_date_from_date_str_with_default_value(  input_date):

            try:
                
                return  str(datetime.strptime(input_date, "%Y-%m-%d").date())
            except:
                return '1900-01-01'        

 
###################################################################################################################################
# T
###################################################################################################################################



   
    @staticmethod
    def get_datetime_from_date_str_with_default_value(  input_date):
        
        try:
            
            return str(datetime.strptime(input_date+ " 00:00:00","%Y-%m-%d %H:%M:%S"))
        except:
            return '1900-01-01 00:00:00'    
        

# ###################################################################################################################################
# # T
# ###################################################################################################################################



   
    @classmethod
    def display_logo(  cls):
        

        img = Image.open('common/ovid_logo.png')
        img.show() 



# ######################################################################################################################################
# This method is exclusively for CHP encounter types.  This function is used to determine the correct ENC_TYPE for the ENCOUNTER TABLE, 
# the DIAGNOSIS TABLE AND the PROCEDURES TABLE. 
# ######################################################################################################################################


    @classmethod
    def get_encounter_type(cls,form_code: Column,type_bill_cd: Column,revenue_cd: Column,hcpcs_cd: Column,
        place_of_service_cd: Column,mod1_cd: Column,mod2_cd: Column,mod3_cd: Column,mod4_cd: Column,) -> Column:
        
        def _ut(c: Column) -> Column:
            return F.upper(F.trim(F.coalesce(c, F.lit(""))))

        def _prefix_in(c: Column, prefixes) -> Column:
            pat = "^(?:" + "|".join(re.escape(p) for p in prefixes) + ")"
            return _ut(c).rlike(pat)

        def _in_set(c: Column, values) -> Column:
            return _ut(c).isin(*values)

        def _numeric_between(c: Column, lo: int, hi: int) -> Column:
            num = F.regexp_replace(F.coalesce(F.trim(c), F.lit("")), r"[^0-9]", "")
            is_num = num.rlike("^[0-9]+$")
            ci = num.cast("int")
            return is_num & (ci >= F.lit(lo)) & (ci <= F.lit(hi))

        def _any_mod_in(mods, vals) -> Column:
            checks = [_in_set(m, vals) for m in mods]
            return reduce(lambda a, b: a | b, checks, F.lit(False))

        fc_u = _in_set(form_code, ["U"])  # Institutional
        fc_h = _in_set(form_code, ["H"])  # Professional

        # Shared cues
        rev_is_ed = _prefix_in(revenue_cd, ["045", "0981"])
        rev_is_os = _in_set(revenue_cd, ["0760", "0761", "0762", "0769"])
        hcpcs_is_ed = _numeric_between(hcpcs_cd, 99281, 99285)
        hcpcs_is_os = _in_set(hcpcs_cd, ["G0244", "G0378", "99217", "99218", "99219", "99220", "99224", "99225", "99226"])

        # Telehealth cues (include G0)
        pos_is_tele = _in_set(place_of_service_cd, ["02", "10"]) | _any_mod_in(
            [mod1_cd, mod2_cd, mod3_cd, mod4_cd], ["GT", "GQ", "95", "G0"]
        )

        # Type of Bill groups (prefix)
        tb_ip = _prefix_in(type_bill_cd, ["11", "12", "41", "84", "011", "012", "041", "084"])
        tb_is = _prefix_in(
            type_bill_cd,
            ["15", "16", "18", "21", "22", "25", "26", "28", "31", "32", "35", "36", "38", "42", "81", "82", "85", "86",
            "015", "016", "018", "021", "022", "025", "026", "028", "031", "032", "035", "036", "038", "042", "081", "082", "085", "086"],
        )
        tb_av = _prefix_in(type_bill_cd, ["13", "14", "23", "24", "33", "34", "7", "83", "013", "014", "023", "024", "033", "034", "07", "083"])
        tb_nonblank = F.length(_ut(type_bill_cd)) > 0

        # Professional cues
        pos_is_ic = _in_set(place_of_service_cd, ["21", "23", "25"])
        hcpcs_is_ic_consult = _numeric_between(hcpcs_cd, 99251, 99255)

        # Institutional path (use your required precedence: IE/EI > IP > IS > ED > OS > AV > TH > OT)
        expr = (
            F.when(fc_u & tb_ip & (rev_is_ed | hcpcs_is_ed), F.lit("EI"))
            .when(fc_u & tb_ip, F.lit("IP"))
            .when(fc_u & tb_is, F.lit("IS"))
            .when(fc_u & (rev_is_ed | hcpcs_is_ed), F.lit("ED"))
            .when(fc_u & (rev_is_os | hcpcs_is_os), F.lit("OS"))
            .when(fc_u & tb_av, F.lit("AV"))
            .when(fc_u & pos_is_tele, F.lit("TH"))
            .when(fc_u & tb_nonblank, F.lit("OT"))
        )

        # Professional path (precedence: ED > IC > OS > TH > AV > OT)
        expr = (
            expr
            .when(fc_h & hcpcs_is_ed, F.lit("ED"))
            .when(fc_h & (pos_is_ic | hcpcs_is_ic_consult), F.lit("IC"))
            .when(fc_h & hcpcs_is_os, F.lit("OS"))
            .when(fc_h & pos_is_tele, F.lit("TH"))
            .when(fc_h & (_in_set(place_of_service_cd, ["03","05","07","09","11","12","13","14","15","16","17","18","19","20","22","33","49","50","71","72"])
                        | _in_set(hcpcs_cd, ["99202","99203","99204","99205","99211","99212","99213","99214","99215","99241","99242","99243","99244","99245",
                                            "99341","99342","99343","99344","99345","99347","99348","99349","99350","99381","99382","99383","99384","99385",
                                            "99386","99387","99391","99392","99393","99394","99395","99396","99397","99401","99402","99403","99404","99411",
                                            "99412","99429","99455","99456","99483","G0402","G0438","G0439","G0463","T1015"])),
                F.lit("AV"))
            .when(
                fc_h & _in_set(
                    place_of_service_cd,
                    ["", " ", "01", "1", "10", "12", "13", "14", "15", "16", "27", "28", "29", "30", "31", "32",
                    "34", "41", "42", "51", "52", "54", "55", "56", "61", "81", "99", "XX", "FI"],
                ),
                F.lit("OT"),
            )
            .when(fc_h, F.lit("AV"))  # final fallback for H
        )

        return expr.otherwise(F.lit("NI"))

